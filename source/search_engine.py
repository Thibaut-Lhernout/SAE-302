"""
Search engine module for the SAE32 application.

- Parses TXT, HTML, PDF and XLSX files.
- Implements search logic including regex and basic Boolean operators (AND / OR).
"""

from __future__ import annotations

import os
import re
from typing import Any, Iterable, List

import openpyxl
from bs4 import BeautifulSoup
from pypdf import PdfReader

# --- Constants ---

# Relative path to documents folder from this script
DOCUMENTS_DIR: str = os.path.abspath(
    os.path.join(os.path.dirname(__file__), "..", "documents")
)

# Number of characters to show around the match
CONTEXT_LENGTH: int = 80


def get_context(text: str, match_index: int, length: int) -> str:
    """
    Extract a substring around the found match for context.

    Parameters
    ----------
    text : str
        Full text line or cell content from which to extract context.
    match_index : int
        Index of the match inside the text.
    length : int
        Maximum length of the context string around the match.

    Returns
    -------
    str
        A shortened string centered around the match, with ellipsis markers.
    """
    start = max(0, match_index - length // 2)
    end = min(len(text), match_index + length // 2)
    snippet = text[start:end].replace("\n", " ").strip()
    return f"...{snippet}..."


def search_txt(filepath: str, pattern: str, is_regex: bool) -> List[dict[str, Any]]:
    """
    Parse and search within a TXT file.
    """
    results: List[dict[str, Any]] = []

    try:
        with open(filepath, "r", encoding="utf-8", errors="ignore") as file:
            for line_idx, line in enumerate(file, 1):
                match = re.search(pattern, line) if is_regex else (pattern in line)
                if not match:
                    continue

                idx = match.start() if is_regex else line.find(pattern)
                results.append(
                    {
                        "file": os.path.basename(filepath),
                        "type": "txt",
                        "location": f"Line {line_idx}",
                        "context": get_context(line, idx, CONTEXT_LENGTH),
                    }
                )
    except Exception as error:
        print(f"[ERROR] Could not read TXT {filepath}: {error}")

    return results


def search_html(filepath: str, pattern: str, is_regex: bool) -> List[dict[str, Any]]:
    """
    Parse and search within an HTML file using BeautifulSoup.
    """
    results: List[dict[str, Any]] = []

    try:
        with open(filepath, "r", encoding="utf-8", errors="ignore") as file:
            soup = BeautifulSoup(file, "html.parser")
            lines = soup.get_text(separator="\n").splitlines()

        for line_idx, line in enumerate(lines, 1):
            if not line.strip():
                continue

            match = re.search(pattern, line) if is_regex else (pattern in line)
            if not match:
                continue

            idx = match.start() if is_regex else line.find(pattern)
            results.append(
                {
                    "file": os.path.basename(filepath),
                    "type": "html",
                    "location": f"Line {line_idx} (Rendered)",
                    "context": get_context(line, idx, CONTEXT_LENGTH),
                }
            )
    except Exception as error:
        print(f"[ERROR] Could not read HTML {filepath}: {error}")

    return results


def search_pdf(filepath: str, pattern: str, is_regex: bool) -> List[dict[str, Any]]:
    """
    Parse and search within a PDF file using pypdf.
    """
    results: List[dict[str, Any]] = []

    try:
        reader = PdfReader(filepath)

        for page_idx, page in enumerate(reader.pages, 1):
            text = page.extract_text()
            if not text:
                continue

            for line in text.splitlines():
                match = re.search(pattern, line) if is_regex else (pattern in line)
                if not match:
                    continue

                idx = match.start() if is_regex else line.find(pattern)
                results.append(
                    {
                        "file": os.path.basename(filepath),
                        "type": "pdf",
                        "location": f"Page {page_idx}",
                        "context": get_context(line, idx, CONTEXT_LENGTH),
                    }
                )
    except Exception as error:
        print(f"[ERROR] Could not read PDF {filepath}: {error}")

    return results


def search_excel(filepath: str, pattern: str, is_regex: bool) -> List[dict[str, Any]]:
    """
    Parse and search within an Excel file using openpyxl.
    """
    results: List[dict[str, Any]] = []

    try:
        workbook = openpyxl.load_workbook(filepath, data_only=True)

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue

                    cell_text = str(cell.value)
                    match = (
                        re.search(pattern, cell_text)
                        if is_regex
                        else (pattern in cell_text)
                    )
                    if not match:
                        continue

                    idx = match.start() if is_regex else cell_text.find(pattern)
                    results.append(
                        {
                            "file": os.path.basename(filepath),
                            "type": "xlsx",
                            "location": f"Sheet '{sheet_name}' | Cell {cell.coordinate}",
                            "context": get_context(cell_text, idx, CONTEXT_LENGTH),
                        }
                    )
    except Exception as error:
        print(f"[ERROR] Could not read Excel {filepath}: {error}")

    return results


def _iter_supported_files(directory: str) -> Iterable[tuple[str, str]]:
    """
    Yield (filepath, extension) pairs for all files in the documents directory.
    """
    for filename in os.listdir(directory):
        filepath = os.path.join(directory, filename)
        _, ext = os.path.splitext(filename)
        yield filepath, ext.lower()


def process_search(
    query: str, allowed_extensions: list[str], use_regex: bool
) -> List[dict[str, Any]]:
    """
    Main search entry point. Dispatches to file handlers.
    """
    all_results: List[dict[str, Any]] = []

    # --- Basic Boolean Logic (AND / OR) when not using regex ---
    search_terms: list[str] = [query]
    operator: str | None = None

    if not use_regex:
        if " AND " in query:
            search_terms = [t.strip() for t in query.split(" AND ") if t.strip()]
            operator = "AND"
        elif " OR " in query:
            search_terms = [t.strip() for t in query.split(" OR ") if t.strip()]
            operator = "OR"

    if not os.path.exists(DOCUMENTS_DIR):
        print(f"[WARNING] Directory not found: {DOCUMENTS_DIR}")
        return []

    for filepath, ext in _iter_supported_files(DOCUMENTS_DIR):
        if allowed_extensions and ext not in allowed_extensions:
            continue

        handler = None
        if ext == ".txt":
            handler = search_txt
        elif ext == ".html":
            handler = search_html
        elif ext == ".pdf":
            handler = search_pdf
        elif ext in [".xlsx", ".xls"]:
            handler = search_excel

        if handler is None:
            continue

        # Collect hits for this file, term by term
        per_term_hits: list[List[dict[str, Any]]] = []

        for term in search_terms:
            hits = handler(filepath, term, use_regex)
            per_term_hits.append(hits)

        if operator == "AND":
            # Keep only files that match all terms
            if all(per_term_hits):
                for hits in per_term_hits:
                    all_results.extend(hits)
        else:
            # OR or no operator: merge all hits
            for hits in per_term_hits:
                all_results.extend(hits)

    return all_results
